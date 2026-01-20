"""
Voice Count Excel Export - Generate Excel reports from voice counting sessions.

This module handles exporting voice count sessions to Excel format with:
- Results ordered by inventory sheet template
- Variance analysis compared to system inventory
- Full transcript log
- Session statistics
"""

from io import BytesIO
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def get_inventory_order_from_template(template_file) -> List[str]:
    """
    Extract item order from uploaded Excel template.

    Args:
        template_file: Uploaded Excel file (UploadedFile from Streamlit)

    Returns:
        List of item_ids in the order they appear in the template
    """
    try:
        # Read the first sheet
        df = pd.read_excel(template_file, sheet_name=0, skiprows=4)

        # Look for Item column (should be first column based on BEVWEEKLY format)
        if 'Item' in df.columns or len(df.columns) > 0:
            item_col = 'Item' if 'Item' in df.columns else df.columns[0]
            items = df[item_col].dropna().astype(str).str.strip().tolist()

            # Filter out TOTAL rows and empty strings
            items = [i for i in items if i and not i.upper().startswith('TOTAL')]

            return items
        else:
            return []
    except Exception as e:
        print(f"Error reading template: {e}")
        return []


def get_default_inventory_order(inventory_dataset, inventory_layout: Optional[Dict] = None) -> List[str]:
    """
    Get default inventory order based on location groups or category.

    Args:
        inventory_dataset: InventoryDataset object
        inventory_layout: Optional dict from inventory_layout.json

    Returns:
        List of item_ids in default order
    """
    items_list = []

    if inventory_layout:
        # Order by physical location groups
        location_order = [
            "well 1", "well 2", "speed rack 1", "top shelf",
            "bottom cabinet", "zipparita table", "bottle fridge",
            "reach in", "chilled fridge"
        ]

        # Group items by location
        for location in location_order:
            location_items = [
                item_id for item_id, item in inventory_dataset.items.items()
                if item.location == location
            ]
            items_list.extend(sorted(location_items))

        # Add items without location
        no_location_items = [
            item_id for item_id, item in inventory_dataset.items.items()
            if not item.location or item.location not in location_order
        ]
        items_list.extend(sorted(no_location_items))
    else:
        # Order by category, then vendor, then name
        category_order = [
            "Whiskey", "Bourbon", "Scotch", "Vodka", "Gin", "Rum",
            "Tequila", "Liqueur", "Well", "Wine", "Draft Beer",
            "Bottled Beer", "Juice", "Bar Consumables"
        ]

        sorted_items = sorted(
            inventory_dataset.items.items(),
            key=lambda x: (
                category_order.index(x[1].category) if x[1].category in category_order else 999,
                x[1].vendor,
                x[1].display_name
            )
        )

        items_list = [item_id for item_id, _ in sorted_items]

    return items_list


def export_voice_count_to_excel(
    session,
    inventory_dataset,
    inventory_order: Optional[List[str]] = None,
    system_inventory: Optional[pd.DataFrame] = None
) -> BytesIO:
    """
    Create Excel workbook with voice count results.

    Args:
        session: VoiceCountSession object
        inventory_dataset: InventoryDataset for item details
        inventory_order: Optional list of item_ids for custom ordering
        system_inventory: Optional DataFrame with system inventory for variance

    Returns:
        BytesIO buffer with Excel file
    """
    wb = Workbook()

    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    Font(bold=True)

    high_conf_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
    med_conf_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Light yellow
    low_conf_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Light red

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Count Results (in inventory order)
    ws1 = wb.active
    ws1.title = "Voice Count"

    # Add title and metadata
    ws1['A1'] = f"HoundCOGS Voice Count - {session.session_name}"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A2'] = f"Date: {session.created_at.strftime('%Y-%m-%d %H:%M')}"
    ws1['A3'] = f"Status: {session.status.upper()}"

    # Headers (row 5)
    headers = ['Item Name', 'Category', 'Vendor', 'Location', 'Count', 'System', 'Variance', 'Confidence', 'Notes']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=5, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Create a map of item_id -> voice count record
    count_map = {}
    for record in session.records:
        if record.matched_item_id and record.is_verified:
            count_map[record.matched_item_id] = record

    # If no custom order provided, use all matched items
    if not inventory_order:
        inventory_order = sorted(count_map.keys())

    # Populate data rows
    row = 6
    total_variance_value = 0

    for item_id in inventory_order:
        if item_id not in inventory_dataset.items:
            continue

        item = inventory_dataset.items[item_id]
        record = count_map.get(item_id)

        if not record:
            continue  # Skip items not counted

        # Get system inventory if available
        system_qty = None
        variance = None
        variance_value = None

        if system_inventory is not None and not system_inventory.empty:
            system_row = system_inventory[system_inventory['item_id'] == item_id]
            if not system_row.empty:
                system_qty = system_row.iloc[0]['on_hand']
                if record.count_value is not None:
                    variance = record.count_value - system_qty
                    if item.unit_cost:
                        variance_value = variance * item.unit_cost
                        total_variance_value += variance_value

        # Write row data
        ws1.cell(row=row, column=1, value=item.display_name).border = thin_border
        ws1.cell(row=row, column=2, value=item.category).border = thin_border
        ws1.cell(row=row, column=3, value=item.vendor).border = thin_border
        ws1.cell(row=row, column=4, value=item.location or "").border = thin_border
        ws1.cell(row=row, column=5, value=record.count_value).border = thin_border
        ws1.cell(row=row, column=6, value=system_qty).border = thin_border
        ws1.cell(row=row, column=7, value=variance).border = thin_border

        # Confidence with color coding
        conf_cell = ws1.cell(row=row, column=8, value=f"{record.confidence_score:.0%}")
        conf_cell.border = thin_border
        if record.confidence_score >= 0.85:
            conf_cell.fill = high_conf_fill
        elif record.confidence_score >= 0.70:
            conf_cell.fill = med_conf_fill
        else:
            conf_cell.fill = low_conf_fill

        ws1.cell(row=row, column=9, value=record.notes or "").border = thin_border

        row += 1

    # Auto-size columns
    for col in range(1, 10):
        ws1.column_dimensions[get_column_letter(col)].width = 15

    # Sheet 2: Variance Analysis
    ws2 = wb.create_sheet("Variance")
    ws2['A1'] = "Variance Analysis"
    ws2['A1'].font = Font(bold=True, size=14)

    variance_headers = ['Item Name', 'Voice Count', 'System Inventory', 'Variance', 'Variance %', 'Value Impact']
    for col, header in enumerate(variance_headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Only show items with variance
    row = 4
    for item_id in inventory_order:
        if item_id not in count_map or item_id not in inventory_dataset.items:
            continue

        item = inventory_dataset.items[item_id]
        record = count_map[item_id]

        if system_inventory is not None and not system_inventory.empty and record.count_value is not None:
            system_row = system_inventory[system_inventory['item_id'] == item_id]
            if not system_row.empty:
                system_qty = system_row.iloc[0]['on_hand']
                variance = record.count_value - system_qty

                # Only show if there's a variance
                if abs(variance) > 0.01:
                    variance_pct = (variance / system_qty * 100) if system_qty != 0 else 0
                    variance_value = variance * item.unit_cost if item.unit_cost else 0

                    ws2.cell(row=row, column=1, value=item.display_name)
                    ws2.cell(row=row, column=2, value=record.count_value)
                    ws2.cell(row=row, column=3, value=system_qty)
                    ws2.cell(row=row, column=4, value=variance)
                    ws2.cell(row=row, column=5, value=f"{variance_pct:.1f}%")
                    ws2.cell(row=row, column=6, value=f"${variance_value:.2f}" if variance_value else "")

                    row += 1

    # Summary row
    if row > 4:
        ws2.cell(row=row + 1, column=1, value="TOTAL VARIANCE:").font = Font(bold=True)
        ws2.cell(row=row + 1, column=6, value=f"${total_variance_value:.2f}").font = Font(bold=True)

    # Auto-size columns
    for col in range(1, 7):
        ws2.column_dimensions[get_column_letter(col)].width = 15

    # Sheet 3: Transcript Log
    ws3 = wb.create_sheet("Transcript Log")
    ws3['A1'] = "Full Transcript Log"
    ws3['A1'].font = Font(bold=True, size=14)

    log_headers = ['Time', 'Raw Transcript', 'Cleaned Transcript', 'Matched Item', 'Count', 'Confidence', 'Verified']
    for col, header in enumerate(log_headers, 1):
        cell = ws3.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # All records in chronological order
    row = 4
    for record in sorted(session.records, key=lambda r: r.timestamp):
        item_name = ""
        if record.matched_item_id and record.matched_item_id in inventory_dataset.items:
            item_name = inventory_dataset.items[record.matched_item_id].display_name

        ws3.cell(row=row, column=1, value=record.timestamp.strftime('%H:%M:%S'))
        ws3.cell(row=row, column=2, value=record.raw_transcript)
        ws3.cell(row=row, column=3, value=record.cleaned_transcript or record.raw_transcript)
        ws3.cell(row=row, column=4, value=item_name)
        ws3.cell(row=row, column=5, value=record.count_value)
        ws3.cell(row=row, column=6, value=f"{record.confidence_score:.0%}")
        ws3.cell(row=row, column=7, value="✓" if record.is_verified else "")

        row += 1

    # Auto-size columns
    for col in range(1, 8):
        ws3.column_dimensions[get_column_letter(col)].width = 18

    # Sheet 4: Summary
    ws4 = wb.create_sheet("Summary")
    ws4['A1'] = "Session Summary"
    ws4['A1'].font = Font(bold=True, size=14)

    # Calculate statistics
    total_records = len(session.records)
    verified_records = len([r for r in session.records if r.is_verified])
    matched_records = len([r for r in session.records if r.matched_item_id])
    high_conf_records = len([r for r in session.records if r.confidence_score >= 0.85])
    unmatched_records = len([r for r in session.records if not r.matched_item_id])

    summary_data = [
        ("Session Name:", session.session_name),
        ("Created:", session.created_at.strftime('%Y-%m-%d %H:%M')),
        ("Last Updated:", session.updated_at.strftime('%Y-%m-%d %H:%M')),
        ("Status:", session.status.upper()),
        ("", ""),
        ("Total Records:", total_records),
        ("Verified Records:", verified_records),
        ("Matched Items:", matched_records),
        ("High Confidence:", f"{high_conf_records} ({high_conf_records/total_records*100:.1f}%)" if total_records > 0 else "0"),
        ("Unmatched:", unmatched_records),
        ("", ""),
        ("Total Variance:", f"${total_variance_value:.2f}" if system_inventory is not None else "N/A"),
    ]

    row = 3
    for label, value in summary_data:
        ws4.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws4.cell(row=row, column=2, value=value)
        row += 1

    # Unmatched items list
    if unmatched_records > 0:
        ws4.cell(row=row + 1, column=1, value="Unmatched Transcripts:").font = Font(bold=True, size=12)
        row += 2

        for record in session.records:
            if not record.matched_item_id:
                ws4.cell(row=row, column=1, value=f"• {record.raw_transcript}")
                ws4.cell(row=row, column=2, value=record.timestamp.strftime('%H:%M:%S'))
                row += 1

    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 30

    # Save to BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
